import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    Font, Alignment, Border, Side, PatternFill, numbers
)
from openpyxl.utils import get_column_letter
from copy import copy
import datetime

# ── スタイル定数 ────────────────────────────────────
GRAY_FILL  = PatternFill("solid", fgColor="F2F2F2")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
NO_FILL    = PatternFill(fill_type=None)

THIN   = Side(style="thin")
MEDIUM = Side(style="medium")

def border(top=None, bottom=None, left=None, right=None):
    return Border(top=top, bottom=bottom, left=left, right=right)

def font(size=10, bold=False, name="MS Pゴシック"):
    return Font(name=name, size=size, bold=bold)

def align(h="general", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def set_cell(ws, coord, value=None, fnt=None, aln=None, brd=None, fill=None, num_fmt=None):
    cell = ws[coord]
    if value is not None:
        cell.value = value
    if fnt:
        cell.font = fnt
    if aln:
        cell.alignment = aln
    if brd:
        cell.border = brd
    if fill:
        cell.fill = fill
    if num_fmt:
        cell.number_format = num_fmt
    return cell


def apply_border_range(ws, min_row, max_row, min_col, max_col,
                        top=None, bottom=None, left=None, right=None,
                        inner_h=None, inner_v=None):
    """範囲にボーダーを適用"""
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            cell = ws.cell(row=r, column=c)
            t = top    if r == min_row  else inner_h
            b = bottom if r == max_row  else inner_h
            l = left   if c == min_col  else inner_v
            rr = right if c == max_col  else inner_v
            cell.border = Border(top=t, bottom=b, left=l, right=rr)


# ── 共通ヘッダー部分を描画 ────────────────────────────
def build_header(ws, doc_title, company_info):
    """
    見積書・請求書共通のヘッダー部分を構築
    company_info = {
        "title": "御見積書" or "請求書",
        "client": "取引先名",
        "doc_date": "日付文字列",
        "subject": "件名",
        "delivery_date": "受渡期日",
        "delivery_place": "受渡場所",
        "payment_method": "取引方法",
        "staff": "担当者名",
        "total_label": "御見積合計金額" or "御請求合計金額",
        "total_ref": "=H46" など,
        "subtotal_ref": "=H44",
    }
    """

    # 行の高さ設定
    row_heights = {
        1:13.5, 2:30.0, 3:27.0, 4:21.75, 5:10.5,
        6:24.0, 7:13.5, 8:22.5, 9:22.5, 10:22.5,
        11:21.75, 12:22.5, 13:11.25, 14:19.5, 15:14.25,
        16:14.25, 17:21.0,
    }
    for r, h in row_heights.items():
        ws.row_dimensions[r].height = h

    # 品目行の高さ（row18〜row43）
    for r in range(18, 48):
        ws.row_dimensions[r].height = 18.0

    # 列幅設定（元のファイルに合わせる）
    col_widths = {
        "A": 4.88, "B": 9.63, "C": 8.0, "D": 8.0, "E": 8.0,
        "F": 5.88, "G": 5.0,  "H": 11.75, "I": 11.0, "J": 12.38,
    }
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    # ── タイトル行（A3） ──
    ws.merge_cells("A1:J2")
    ws.merge_cells("A3:J3")
    set_cell(ws, "A3", doc_title,
             fnt=Font(name="MS Pゴシック", size=18),
             aln=align("center", "center"))

    # ── 日付（J4） ──
    set_cell(ws, "J4", company_info.get("doc_date", ""),
             fnt=font(10),
             aln=align("right", "center"))

    # ── 取引先（A6〜F6 + 様） ──
    ws.merge_cells("A6:E6")
    set_cell(ws, "A6", company_info.get("client", ""),
             fnt=font(12),
             aln=align("left", "center"),
             brd=border(bottom=MEDIUM))
    # B〜E にも下ボーダー
    for col in ["B6", "C6", "D6", "E6"]:
        ws[col].border = border(bottom=MEDIUM)
    set_cell(ws, "F6", "様",
             fnt=font(12),
             aln=align("left", "center"),
             brd=border(bottom=MEDIUM))

    # G列は結合スペーサー
    ws.merge_cells("G6:G12")

    # ── アースワーク会社情報（H6〜J8） ──
    ws.merge_cells("H6:J8")
    addr_text = (
        "アースワーク株式会社\n"
        "〒130-0005　東京都墨田区東駒形2-3-8\n"
        "Tel.03-5819-4550　Fax.03-5819-4504\n"
        "登録番号：T6010601067902"
    )
    set_cell(ws, "H6", addr_text,
             fnt=font(8),
             aln=align("center", "center", wrap=True))

    # ── 本文（A8） ──
    ws.merge_cells("A8:F8")
    set_cell(ws, "A8", "下記の通りご請求申し上げます。",
             fnt=font(10),
             aln=align("left", "center"))

    # ── 件名（A9） ──
    set_cell(ws, "A9", "件名：",
             fnt=font(10), aln=align("center", "center"),
             brd=border(bottom=THIN))
    for col in ["B9","C9","D9","E9","F9"]:
        ws[col].border = border(bottom=THIN)
    # 件名の値（C9に入力）
    set_cell(ws, "C9", company_info.get("subject", ""),
             fnt=font(10), aln=align("left", "center"),
             brd=border(bottom=THIN))

    # ── 受渡期日（A10） ──
    set_cell(ws, "A10", "受渡期日：",
             fnt=font(10), aln=align("center", "center"),
             brd=border(top=THIN, bottom=THIN))
    set_cell(ws, "C10", company_info.get("delivery_date", "ー"),
             fnt=font(10), aln=align("center", "center"),
             brd=border(top=THIN, bottom=THIN))
    for col in ["B10","D10","E10","F10"]:
        ws[col].border = border(top=THIN, bottom=THIN)
    # H10〜J11 担当住所
    ws.merge_cells("H10:J11")

    # ── 受渡場所（A11） ──
    ws.merge_cells("A11:B11")
    set_cell(ws, "A11", "受渡場所：",
             fnt=font(10), aln=align("center", "center"),
             brd=border(top=THIN, bottom=THIN))
    set_cell(ws, "C11", company_info.get("delivery_place", "ー"),
             fnt=font(10), aln=align("center", "center"),
             brd=border(top=THIN, bottom=THIN))
    for col in ["D11","E11","F11"]:
        ws[col].border = border(top=THIN, bottom=THIN)

    # ── 取引方法（A12） ──
    set_cell(ws, "A12", "取引方法：",
             fnt=font(10), aln=align("center", "center"),
             brd=border(top=THIN, bottom=THIN))
    set_cell(ws, "C12", company_info.get("payment_method", "従来通り"),
             fnt=font(10), aln=align("center", "center"),
             brd=border(top=THIN, bottom=THIN))
    for col in ["B12","D12","E12","F12"]:
        ws[col].border = border(top=THIN, bottom=THIN)
    set_cell(ws, "H12", "担当：",
             fnt=font(10), aln=align("center", "center"),
             brd=border(bottom=THIN))
    set_cell(ws, "I12", company_info.get("staff", "木戸　志朗"),
             fnt=font(10), aln=align("center", "center"),
             brd=border(bottom=THIN))
    ws["J12"].border = border(bottom=THIN)

    # ── 区切り線（A13） ──
    ws.merge_cells("A13:J13")
    for col in range(1, 11):
        ws.cell(row=13, column=col).border = border(bottom=MEDIUM)

    # ── 合計金額ボックス（A14〜B15 と H14〜J15） ──
    ws.merge_cells("A14:B15")
    set_cell(ws, "A14", company_info.get("total_label", "御見積合計金額"),
             fnt=font(10), aln=align("center", "center"),
             fill=GRAY_FILL,
             brd=border(top=MEDIUM, bottom=MEDIUM, right=MEDIUM))
    ws["B14"].border = border(top=MEDIUM, right=MEDIUM)
    ws["A15"].border = border(bottom=MEDIUM)
    ws["B15"].border = border(bottom=MEDIUM, right=MEDIUM)

    ws.merge_cells("C14:C15")
    set_cell(ws, "C14", company_info.get("total_ref", "=H46"),
             fnt=Font(name="MS Pゴシック", size=16),
             aln=align("center", "center"),
             brd=border(top=MEDIUM, bottom=MEDIUM, left=MEDIUM, right=MEDIUM),
             num_fmt='#,##0')
    ws["C15"].border = border(bottom=MEDIUM, left=MEDIUM)

    # D〜G（空白スペース）
    for col in ["D14","E14","F14","G14"]:
        ws[col].border = border(top=MEDIUM)
    ws["G14"].border = border(top=MEDIUM, right=MEDIUM)
    for col in ["D15","E15","F15","G15"]:
        ws[col].border = border(bottom=MEDIUM)
    ws["G15"].border = border(bottom=MEDIUM, right=MEDIUM)

    # 税抜金額: H14=ラベル(GRAY), I14:J14マージ=数式値
    set_cell(ws, "H14", "税抜金額",
             fnt=font(10), aln=align("center", "center"),
             fill=GRAY_FILL,
             brd=border(top=MEDIUM, bottom=MEDIUM, left=MEDIUM))
    set_cell(ws, "H15", None, brd=border(bottom=MEDIUM, left=MEDIUM))
    ws.merge_cells("I14:J14")
    ws.merge_cells("I15:J15")
    set_cell(ws, "I14", company_info.get("subtotal_ref", "=H44"),
             fnt=Font(name="MS Pゴシック", size=12),
             aln=align("center", "center"),
             brd=border(top=MEDIUM, bottom=MEDIUM, left=MEDIUM),
             num_fmt='#,##0')
    ws["I15"].border = border(bottom=MEDIUM, left=MEDIUM)

    # ── row16（空白） ──

    # ── 品目ヘッダー（row17） ──
    ws.merge_cells("A17:E17")
    set_cell(ws, "A17", "品名・項目",
             fnt=font(10), aln=align("center", "center"),
             fill=GRAY_FILL,
             brd=border(top=MEDIUM, bottom=THIN, right=THIN))
    for c in ["B17","C17","D17","E17"]:
        ws[c].fill = GRAY_FILL
        ws[c].border = border(top=MEDIUM, bottom=THIN)
    ws["E17"].border = border(top=MEDIUM, bottom=THIN, right=THIN)

    for col, label in [("F17","数量"),("G17","単位"),("H17","単価"),("I17","金額"),("J17","備考")]:
        set_cell(ws, col, label,
                 fnt=font(10), aln=align("center", "center"),
                 fill=GRAY_FILL,
                 brd=border(top=MEDIUM, bottom=THIN, left=THIN, right=THIN))


# ── 品目行を描画 ────────────────────────────────────
def build_items(ws, items, start_row=18, max_rows=26):
    """
    items = [{"name":"", "qty":1, "unit":"式", "price":0, "note":""}]
    max_rows: 最大品目行数（元のフォーマットは row18〜row43 = 26行）
    """
    for i in range(max_rows):
        r = start_row + i * 2 - (i > 0) * 0  # 元ファイルは1品目=2行のようだが実際は1行
        r = start_row + i
        item = items[i] if i < len(items) else {}
        fill = GRAY_FILL if i % 2 == 0 else NO_FILL

        # A〜E: 品名（マージ）
        ws.merge_cells(f"A{r}:E{r}")
        set_cell(ws, f"A{r}", item.get("name", ""),
                 fnt=font(11), aln=align("left", "center"),
                 fill=fill,
                 brd=border(bottom=THIN, right=THIN))
        for c in ["B","C","D","E"]:
            ws[f"{c}{r}"].fill = fill
            ws[f"{c}{r}"].border = border(bottom=THIN)
        ws[f"E{r}"].border = border(bottom=THIN, right=THIN)

        # F: 数量
        set_cell(ws, f"F{r}", item.get("qty", "") if item else "",
                 fnt=font(11), aln=align("center", "center"),
                 fill=fill,
                 brd=border(bottom=THIN, left=THIN, right=THIN),
                 num_fmt='#,##0.##')

        # G: 単位
        set_cell(ws, f"G{r}", item.get("unit", "") if item else "",
                 fnt=font(11), aln=align("center", "center"),
                 fill=fill,
                 brd=border(bottom=THIN, left=THIN, right=THIN))

        # H: 単価
        set_cell(ws, f"H{r}", item.get("price", "") if item else "",
                 fnt=font(11), aln=align("right", "center"),
                 fill=fill,
                 brd=border(bottom=THIN, left=THIN, right=THIN),
                 num_fmt='#,##0')

        # I: 金額（数式）
        if item.get("name"):
            set_cell(ws, f"I{r}", f"=IF(ISBLANK(H{r}),\"\",F{r}*H{r})",
                     fnt=font(11), aln=align("right", "center"),
                     fill=fill,
                     brd=border(bottom=THIN, left=THIN, right=THIN),
                     num_fmt='#,##0')
        else:
            set_cell(ws, f"I{r}", f"=IF(ISBLANK(H{r}),\"\",F{r}*H{r})",
                     fnt=font(11), aln=align("right", "center"),
                     fill=fill,
                     brd=border(bottom=THIN, left=THIN, right=THIN),
                     num_fmt='#,##0')

        # J: 備考
        set_cell(ws, f"J{r}", item.get("note", "") if item else "",
                 fnt=font(10), aln=align("left", "center", wrap=True),
                 fill=fill,
                 brd=border(bottom=THIN, left=THIN))


# ── 合計行を描画 ────────────────────────────────────
def build_totals(ws, subtotal_row, notes_text="", bank_info=None):
    """
    subtotal_row: 税抜金額の行番号（元フォーマットでは44）
    """
    r = subtotal_row

    # 特記事項 or 振込先
    set_cell(ws, f"A{r}", notes_text or "特記事項",
             fnt=Font(name="MS Pゴシック", size=10, bold=bool(notes_text is None)),
             aln=align("left", "center"))

    # 振込先情報（請求書用）
    if bank_info:
        ws.merge_cells(f"A{r+1}:E{r+2}")
        set_cell(ws, f"A{r+1}", bank_info,
                 fnt=font(9), aln=align("left", "center", wrap=True))

    # 税抜金額
    sum_range = f"I18:I{r-1}"
    set_cell(ws, f"F{r}", "税抜金額",
             fnt=font(10), aln=align("center", "center"),
             brd=border())
    set_cell(ws, f"H{r}", f"=SUM({sum_range})",
             fnt=Font(name="MS Pゴシック", size=12),
             aln=align("center", "center"),
             num_fmt='#,##0')
    ws.merge_cells(f"H{r}:J{r}")

    # 消費税
    set_cell(ws, f"F{r+1}", "消費税",
             fnt=font(10), aln=align("center", "center"))
    set_cell(ws, f"H{r+1}", f"=H{r}*0.1",
             fnt=Font(name="MS Pゴシック", size=12),
             aln=align("center", "center"),
             num_fmt='#,##0')
    ws.merge_cells(f"H{r+1}:J{r+1}")

    # 総額
    set_cell(ws, f"F{r+2}", "総額",
             fnt=font(10), aln=align("center", "center"))
    set_cell(ws, f"H{r+2}", f"=H{r}+H{r+1}",
             fnt=Font(name="MS Pゴシック", size=12),
             aln=align("center", "center"),
             num_fmt='#,##0')
    ws.merge_cells(f"H{r+2}:J{r+2}")

    return r  # subtotal_row を返す（参照用）


# ── 見積書シートを生成 ──────────────────────────────
def create_estimate_sheet(wb, items=None, info=None):
    ws = wb.create_sheet("アースワーク_見積書")
    if items is None:
        items = []
    if info is None:
        info = {}

    # 品目は最大26行（row18〜row43）
    MAX_ITEMS = 26
    subtotal_row = 18 + MAX_ITEMS  # = 44

    company_info = {
        "total_label": "御見積合計金額",
        "total_ref":    f"=H{subtotal_row+2}",
        "subtotal_ref": f"=H{subtotal_row}",
        "client":       info.get("client", ""),
        "doc_date":     info.get("doc_date", ""),
        "subject":      info.get("subject", ""),
        "delivery_date":  info.get("delivery_date", "ー"),
        "delivery_place": info.get("delivery_place", "ー"),
        "payment_method": info.get("payment_method", "従来通り"),
        "staff":          info.get("staff", "木戸　志朗"),
    }

    build_header(ws, "御見積書", company_info)
    build_items(ws, items, start_row=18, max_rows=MAX_ITEMS)
    build_totals(ws, subtotal_row)

    # 印刷設定
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = 9  # A4
    ws.print_area = f"A1:J{subtotal_row+4}"

    return ws


# ── 請求書シートを生成 ──────────────────────────────
def create_invoice_sheet(wb, items=None, info=None):
    ws = wb.create_sheet("アースワーク_請求書")
    if items is None:
        items = []
    if info is None:
        info = {}

    MAX_ITEMS = 26
    subtotal_row = 18 + MAX_ITEMS  # = 44

    company_info = {
        "total_label": "御請求合計金額",
        "total_ref":    f"=H{subtotal_row+2}",
        "subtotal_ref": f"=H{subtotal_row}",
        "client":       info.get("client", ""),
        "doc_date":     info.get("doc_date", ""),
        "subject":      info.get("subject", ""),
        "delivery_date":  info.get("delivery_date", "ー"),
        "delivery_place": info.get("delivery_place", "ー"),
        "payment_method": info.get("payment_method", "従来通り"),
        "staff":          info.get("staff", "木戸　志朗"),
    }

    build_header(ws, "請求書", company_info)
    build_items(ws, items, start_row=18, max_rows=MAX_ITEMS)

    bank_info = (
        "三井住友銀行　トランクNORTH支店（店番：403）\n"
        "普通預金　0260571\n"
        "アースワーク（カ"
    )
    build_totals(ws, subtotal_row, notes_text="振込先", bank_info=bank_info)

    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = 9
    ws.print_area = f"A1:J{subtotal_row+4}"

    return ws


# ── メイン：空白テンプレート＋サンプル入力済みの2ブックを生成 ──
def main():
    today = datetime.date.today().strftime("%Y年%m月%d日")

    # ── ① 空白テンプレート（見積書＋請求書の2シート） ──
    wb_template = Workbook()
    del wb_template["Sheet"]  # デフォルトシートを削除

    create_estimate_sheet(wb_template, items=[], info={
        "doc_date": today,
        "client": "",
        "subject": "",
    })
    create_invoice_sheet(wb_template, items=[], info={
        "doc_date": today,
        "client": "",
        "subject": "",
    })

    wb_template.save("/home/claude/bizdb/アースワーク_書類テンプレート.xlsx")
    print("テンプレート生成完了")

    # ── ② サンプルデータ入り（確認用） ──
    sample_items = [
        {"name": "①芝浦ポスター撤去",          "qty": 1, "unit": "式", "price": 30600, "note": "250171"},
        {"name": "②看板設置工事",               "qty": 1, "unit": "式", "price": 45000, "note": ""},
        {"name": "③デザイン・制作費",           "qty": 1, "unit": "式", "price": 68000, "note": ""},
        {"name": "④施工管理費",                 "qty": 1, "unit": "式", "price": 12000, "note": ""},
    ]
    wb_sample = Workbook()
    del wb_sample["Sheet"]

    create_estimate_sheet(wb_sample, items=sample_items, info={
        "doc_date": today,
        "client": "〇〇株式会社",
        "subject": "芝浦ポスター撤去及び看板設置工事一式",
        "delivery_date": "2025年12月31日",
        "delivery_place": "東京都港区芝浦",
    })

    # 請求書は金額を一部変更（値引き対応デモ）
    invoice_items = [
        {"name": "①芝浦ポスター撤去",          "qty": 1, "unit": "式", "price": 30600, "note": "250171"},
        {"name": "②看板設置工事",               "qty": 1, "unit": "式", "price": 42000, "note": "値引後"},  # ← 変更
        {"name": "③デザイン・制作費",           "qty": 1, "unit": "式", "price": 68000, "note": ""},
        {"name": "④施工管理費",                 "qty": 1, "unit": "式", "price": 12000, "note": ""},
    ]
    create_invoice_sheet(wb_sample, items=invoice_items, info={
        "doc_date": today,
        "client": "〇〇株式会社",
        "subject": "芝浦ポスター撤去及び看板設置工事一式",
        "delivery_date": "2025年12月31日",
    })

    wb_sample.save("/home/claude/bizdb/アースワーク_サンプル.xlsx")
    print("サンプル生成完了")


if __name__ == "__main__":
    main()
